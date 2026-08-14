#!/usr/bin/env python3
# Pipeline Colombia v2: histórico (fuente vieja) + ventana móvil de 4 semanas
# (data nueva del stakeholder: rt_data_v2.csv). Polígono = coverageCurrent.
# Genera const D (11 semanas) con: b,k,c,storeId,cc,fin[],rt[]  e inyecta en index.html.
#
# Lógica del stakeholder (evaluada en JS al render):
#   avg4w = promedio RTWT de las últimas 4 semanas (con dato).
#   coverageResult = coverageCurrent  si avg4w<=8  else 0  (pierde cobertura + penalización).
#   estados por avg4w: 🟢<2  🟡2-5  🟠5-8  🔴>8 (pierde)  ⚪ sin datos.
import openpyxl, csv, json, sys, re, datetime, os
from collections import defaultdict

NEW = sys.argv[1] if len(sys.argv) > 1 else 'rt_data_v2.csv'
OLD = 'Turbology.xlsx'

# ---- semanas ----
HIST = ['May 10','May 17','May 24','May 31','Jun 7','Jun 14','Jun 21']  # etiqueta = último día (domingo) de cada semana
HIST_DATES = ['2026-05-04','2026-05-11','2026-05-18','2026-05-25','2026-06-01','2026-06-08','2026-06-15']  # fechas reales (lunes) para el match
NEW_WEEKS = ['Jun 28','Jul 5','Jul 12','Jul 19','Jul 26','Aug 2','Aug 9']  # último día de cada semana (Jun22..Aug03 -> +6d)
WEEKS = HIST + NEW_WEEKS                                               # 12 semanas
NW = len(WEEKS)
N4 = 4                                                                 # ventana móvil fija de 4 semanas (Jun 29–Jul 20)

def to_m(v):
    if v in (None, ''): return None
    v = float(v)
    return v*1000 if v < 50 else v          # km -> m (1.0, 2.4 ...)
MAY11 = {1.0:1000.0, 2.0:2400.0, 3.0:3000.0}
def fin_m(dd, v):
    if dd == '2026-05-11':
        if v in (None, ''): return None
        return MAY11.get(float(v), to_m(v))
    return to_m(v)
def rtclean(v):
    if v in (None, ''): return None
    v = float(v)
    return None if v > 60 else v            # descarta seriales de fecha / basura

# ---- 0) mapeo cocina -> ops (del archivo KDS) ----
coc2op = json.load(open('cocina2ops.json')) if os.path.exists('cocina2ops.json') else {}

# ---- 1) historia (weeks 0..6) + ciudad, desde la fuente vieja ----
s2c = json.load(open('store2cocina.json'))            # 'CO...' -> cocina
wb = openpyxl.load_workbook(OLD, read_only=True, data_only=True)
ws = wb['Raw']
H = {c:i for i,c in enumerate(next(ws.iter_rows(min_row=1,max_row=1,values_only=True)))}
hist_rt = defaultdict(lambda: [None]*NW)
hist_fin = defaultdict(lambda: [None]*NW)
city_of = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    sid = r[H['Store ID']]
    if not sid: continue
    d = r[H['Semana']]
    dd = (d if isinstance(d,datetime.datetime) else datetime.datetime.strptime(str(d),'%d/%m/%Y')).strftime('%Y-%m-%d')
    key = re.sub(r'\D','',str(sid))             # id numérico
    if r[H['City Name']]: city_of[key] = r[H['City Name']]   # ciudad de cualquier semana
    if dd not in HIST_DATES: continue           # rt/fin solo de semanas históricas
    w = HIST_DATES.index(dd)
    hist_rt[key][w] = rtclean(r[H['Avg. RTWT']])
    hist_fin[key][w] = fin_m(dd, r[H['Final Size']])

# ---- 2) data nueva (weeks 7..10) ----
def numf(x):
    try: return float(x)
    except: return None
# ---- data del dashboard nuevo (29-jun..20-jul + coberturas actualizadas) ----
UP = 'foodology_rt_dashboard.html'
up = {}
if os.path.exists(UP):
    _h = open(UP, encoding='utf-8', errors='replace').read()
    for _m in re.finditer(r'"storeId"', _h):
        _i = _m.start(); _s = _h.rfind('{', 0, _i); _d = 0; _e = None
        for _j in range(_s, len(_h)):
            if _h[_j] == '{': _d += 1
            elif _h[_j] == '}':
                _d -= 1
                if _d == 0: _e = _j + 1; break
        try: _o = json.loads(_h[_s:_e])
        except: continue
        if 'w4' in _o and 'storeId' in _o:
            up[re.sub(r'\D','',str(_o['storeId']))] = _o
    print('dashboard nuevo: %d tiendas (29-jun..20-jul)' % len(up))

# ---- dashboard más reciente (06-jul..27-jul, SOLO RTWT; sin coberturas) ----
UP2 = 'foodology_rt_dashboard_v2.html'
up2 = {}
if os.path.exists(UP2):
    _h2 = open(UP2, encoding='utf-8', errors='replace').read()
    for _m in re.finditer(r'"storeId"', _h2):
        _i = _m.start(); _s = _h2.rfind('{', 0, _i); _d = 0; _e = None
        for _j in range(_s, len(_h2)):
            if _h2[_j] == '{': _d += 1
            elif _h2[_j] == '}':
                _d -= 1
                if _d == 0: _e = _j + 1; break
        try: _o = json.loads(_h2[_s:_e])
        except: continue
        if 'w4' in _o and 'storeId' in _o:
            up2[re.sub(r'\D','',str(_o['storeId']))] = _o
    print('dashboard v2: %d tiendas (06-jul..27-jul, RTWT)' % len(up2))

# ---- dashboard más reciente (13-jul..03-ago, RTWT + coverageCurrent) ----
UP3 = 'foodology_rt_dashboard_v3.html'
up3 = {}
if os.path.exists(UP3):
    _h3 = open(UP3, encoding='utf-8', errors='replace').read()
    for _m in re.finditer(r'"storeId"', _h3):
        _i = _m.start(); _s = _h3.rfind('{', 0, _i); _d = 0; _e = None
        for _j in range(_s, len(_h3)):
            if _h3[_j] == '{': _d += 1
            elif _h3[_j] == '}':
                _d -= 1
                if _d == 0: _e = _j + 1; break
        try: _o = json.loads(_h3[_s:_e])
        except: continue
        if 'w4' in _o and 'storeId' in _o:
            up3[re.sub(r'\D','',str(_o['storeId']))] = _o
    print('dashboard v3: %d tiendas (13-jul..03-ago, RTWT+cobertura)' % len(up3))

stores = {}
with open(NEW, encoding='utf-8') as f:
    for row in csv.DictReader(f):
        key = re.sub(r'\D','',str(row['storeId']))
        cc_old = row.get('coverageCurrent')
        cc_old = to_m(cc_old) if cc_old not in (None,'') else None
        b = re.sub(r'\s*-\s*Turbo\s*$','', row['brand'] or '').strip()
        coc = s2c.get('CO'+key) or s2c.get(key) or '—'
        rt = list(hist_rt[key])                 # copia historia (0..6)
        fin = list(hist_fin[key])
        # semana 7 = Jun 22 desde rt_data (w1); polígono = coverageCurrent viejo
        rt[7] = rtclean(row.get('w1')); fin[7] = cc_old
        u = up.get(key)      # 29-jun..20-jul + coverageCurrent
        u2 = up2.get(key)    # 06-jul..27-jul (solo RTWT)
        u3 = up3.get(key)    # 13-jul..03-ago (RTWT + coverageCurrent) — el más reciente
        # cobertura: del más reciente que la traiga (v3 > viejo); si no, cc_old
        _c = (u3.get('coverageCurrent') if u3 else None) or (u.get('coverageCurrent') if u else None)
        cc = to_m(_c) if _c not in (None,'') else cc_old
        ow = [u.get('w1'), u.get('w2'), u.get('w3'), u.get('w4')] if u else [None]*4      # 29jun,06jul,13jul,20jul
        nw = [u2.get('w1'), u2.get('w2'), u2.get('w3'), u2.get('w4')] if u2 else [None]*4   # 06jul,13jul,20jul,27jul
        tw = [u3.get('w1'), u3.get('w2'), u3.get('w3'), u3.get('w4')] if u3 else [None]*4   # 13jul,20jul,27jul,03ago
        pick = lambda *xs: next((x for x in xs if x is not None), None)
        rt[8]  = rtclean(ow[0])                                                   # 29-jun (Jul 5)
        rt[9]  = rtclean(pick(nw[0], ow[1]))                                      # 06-jul (Jul 12)
        rt[10] = rtclean(pick(tw[0], nw[1], ow[2]))                              # 13-jul (Jul 19)
        rt[11] = rtclean(pick(tw[1], nw[2], ow[3]))                              # 20-jul (Jul 26)
        rt[12] = rtclean(pick(tw[2], nw[3]))                                     # 27-jul (Aug 2)
        rt[13] = rtclean(tw[3])                                                   # 03-ago (Aug 9) — nueva semana
        for i in range(8, 14): fin[i] = cc                                        # polígono actual = coverageCurrent (fijo)
        stores[key] = dict(b=b, k=coc, c=city_of.get(key,'—'), sid=key, cc=cc, fin=fin, rt=rt, op=coc2op.get(coc))

D = dict(weeks=WEEKS, n4=N4, stores=list(stores.values()))
json.dump(D, open('D_v2.json','w'), ensure_ascii=False)

# ---- reporte rápido ----
def avg(a):
    a=[x for x in a if x is not None]; return sum(a)/len(a) if a else None
last4 = lambda s: avg(s['rt'][10:14])
allavg = avg([last4(s) for s in stores.values() if last4(s) is not None])
lose = sum(1 for s in stores.values() if (last4(s) or 0) > 8)
covres = [0 if (last4(s) or 0) > 8 else s['cc'] for s in stores.values() if s['cc'] is not None]
print('CO v2 | semanas:', NW, WEEKS)
print('tiendas:', len(stores))
print('Avg RTWT 4 sem (red):', round(allavg,2))
print('tiendas que pierden cobertura (>8):', lose)
print('cobertura red (prom coverageResult/3000):', round(100*avg(covres)/3000,1),'%')
from collections import Counter
print('coverageCurrent dist:', Counter(s['cc'] for s in stores.values()))

if '--inject' in sys.argv:
    idx = open('index_co.html', encoding='utf-8').read()
    payload = 'const D=' + json.dumps(D, ensure_ascii=False) + ';'
    new = re.sub(r'const D=\{.*?\};', lambda m: payload, idx, count=1, flags=re.S)
    assert new != idx and 'const D={' in new
    open('index_co.html','w',encoding='utf-8').write(new)
    print('const D (v2) inyectado en index.html')
