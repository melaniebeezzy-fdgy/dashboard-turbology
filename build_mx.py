#!/usr/bin/env python3
# Pipeline dashboard MÉXICO. Comparativo de 2 semanas (sin histórico largo).
#  - RTWT: actual (Avg. RTWT) vs semana pasada (LW_RTWT)  [FOODOLOGY.xlsx]
#  - Polígono/cobertura: actual (FINAL SIZE) vs propuesto (Size Proposal)  [mx_polygon_proposal]
#  - Ventas: KDS_ventas_mx.xlsx (última semana vs anterior, WoW) + columna ops
#  - Filtro por ciudad y por ops (excluyentes, como Colombia)
# Genera mx_data.json e inyecta const MX en index_mx.html.
import openpyxl, json, sys, re, unicodedata, datetime, os, warnings
from collections import Counter, defaultdict
warnings.filterwarnings('ignore')

DETALLE = sys.argv[1] if len(sys.argv) > 1 else 'FOODOLOGY.xlsx'
PROP = 'mx_polygon_proposal.xlsx'
KDS = 'KDS_ventas_mx.xlsx'
RTWTF = 'MX_rtwt.xlsx'
LV = [1000, 2100, 2400, 2700, 3000]
IDEAL = 3000

def num(x): return float(x) if isinstance(x, (int, float)) else None
def to_m(x):
    v = num(x)
    return None if v is None else (v*1000 if v < 50 else v)
def sac(s): return ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
def avg(a):
    a = [x for x in a if x is not None]; return sum(a)/len(a) if a else None
def lvl(m): return min(LV, key=lambda l: abs(l-m)) if m is not None else None
CITY_MAP = {'Ciudad de México': 'CDMX', 'Merida': 'Mérida'}
# override de ciudad por cocina (para unificar cocinas que vienen con ciudad distinta entre fuentes)
CITY_OVERRIDE = {'Saltillo Centro': 'CDMX'}
def ncity(c): return CITY_MAP.get(str(c), str(c))
def city_for(coc, raw): return CITY_OVERRIDE.get(coc, ncity(raw))

s2c = json.load(open('store2cocina_mx.json'))
coc2op = json.load(open('cocina2ops_mx.json')) if os.path.exists('cocina2ops_mx.json') else {}

# ---- polígono propuesto por tienda ----
prop_of = {}
if os.path.exists(PROP):
    wp = openpyxl.load_workbook(PROP, read_only=True, data_only=True)
    ws = wp[wp.sheetnames[0]]
    Hp = {c: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}
    si, pi = Hp.get('Store ID'), Hp.get('Size Proposal')
    for r in ws.iter_rows(min_row=2, values_only=True):
        if si is None or si >= len(r) or not r[si]: continue
        prop_of[str(r[si])] = to_m(r[pi]) if (pi is not None and pi < len(r)) else None

# ---- tiendas (FOODOLOGY: RTWT actual + LW, polígono actual) ----
wb = openpyxl.load_workbook(DETALLE, read_only=True, data_only=True)
ws = wb['DETALLE']
H = {c: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}
def col(r, n):
    i = H.get(n); return r[i] if i is not None and i < len(r) else None
def stripbrand(s): return re.sub(r'\s*-\s*Turbo\s*$', '', str(s or ''), flags=re.I).strip()
stores = []
for r in ws.iter_rows(min_row=2, values_only=True):
    sid = col(r, 'Store ID')
    if not sid: continue
    coc = s2c.get(str(sid), '—')
    stores.append(dict(sid=str(sid), b=stripbrand(col(r, 'Brand Name')), k=coc,
        c=city_for(coc, col(r, 'City Name') or '—'), op=coc2op.get(coc),
        rt=num(col(r, 'Avg. RTWT')), rtlw=num(col(r, 'LW_RTWT')),
        cur=to_m(col(r, 'Current Size')),
        fin=to_m(col(r, 'FINAL SIZE')), prop=prop_of.get(str(sid))))

# ---- ventas KDS (WoW) ----
def norm(s):
    if s is None: return ''
    s = sac(s).lower().replace('&', ' ').replace('.', ' ').replace('-', ' ')
    drop = {'turbo', 'fdl', 'fd', 'court', 'sandwiches', 'pase', 'desayunos', 'desayuno'}
    return ' '.join(t for t in re.sub(r'[^a-z0-9 ]', ' ', s).split() if t not in drop).strip()
def kit2coc(kid):
    nb = sac(re.sub(r'^\s*\d+\s*', '', str(kid))).lower().strip()
    for c in set(s2c.values()):
        cc = sac(c).lower()
        if cc == nb or nb in cc or cc.split(' /')[0] == nb: return c
    return re.sub(r'^\s*\d+\s*', '', str(kid)).title()
vrecs = []
if os.path.exists(KDS):
    wk = openpyxl.load_workbook(KDS, read_only=True, data_only=True)
    wv = wk['Export']
    Hv = {c: i for i, c in enumerate(next(wv.iter_rows(min_row=1, max_row=1, values_only=True)))}
    def vc(r, n):
        i = Hv.get(n); return r[i] if i is not None and i < len(r) else None
    for r in wv.iter_rows(min_row=2, values_only=True):
        d = vc(r, 'date')
        if not isinstance(d, datetime.datetime): continue
        o = num(vc(r, 'Total orders')) or 0
        _c = kit2coc(vc(r, 'kitchen_id'))
        vrecs.append((d.date(), city_for(_c, vc(r, 'city')), _c,
                      str(vc(r, 'brand') or ''), o, vc(r, 'ops')))
days = [x[0] for x in vrecs]
lastday = max(days) if days else None
sales_week = prevlabel = None
lw_all = pw_all = []
if lastday:
    firstlw = lastday - datetime.timedelta(days=6)
    pfirst = firstlw - datetime.timedelta(days=7); plast = firstlw - datetime.timedelta(days=1)
    sales_week = f"{firstlw.strftime('%d/%m')}–{lastday.strftime('%d/%m/%Y')}"
    prevlabel = f"{pfirst.strftime('%d/%m')}–{plast.strftime('%d/%m')}"
    lw_all = [x for x in vrecs if firstlw <= x[0] <= lastday]
    pw_all = [x for x in vrecs if pfirst <= x[0] <= plast]
bdisp = {}
for x in lw_all + pw_all:
    b = x[3]; bdisp.setdefault(norm(b), b.title())

# ---- RTWT semanal (KDS time mx: diario -> semanal por lunes) ----
rtrows = []   # (wi, city, coc, op, rtwt)
MXW = []
if os.path.exists(RTWTF):
    wr = openpyxl.load_workbook(RTWTF, read_only=True, data_only=True)
    wsr = wr['Export']
    Hr = {c: i for i, c in enumerate(next(wsr.iter_rows(min_row=1, max_row=1, values_only=True)))}
    def rc(r, n):
        i = Hr.get(n); return r[i] if i is not None and i < len(r) else None
    raw = []; mondays = set()
    for r in wsr.iter_rows(min_row=2, values_only=True):
        d = rc(r, 'date')
        if not isinstance(d, datetime.datetime): continue
        rtv = num(rc(r, 'RTWT'))
        if rtv is None: continue
        mon = (d - datetime.timedelta(days=d.weekday())).date()
        mondays.add(mon)
        coc = kit2coc(rc(r, 'kitchen_id'))
        raw.append((mon, city_for(coc, rc(r, 'city')), coc, coc2op.get(coc), rtv))
    ws_sorted = sorted(mondays)
    widx = {m: i for i, m in enumerate(ws_sorted)}
    MXW = [m.strftime('%b %-d') for m in ws_sorted]
    for mon, city, coc, op, rtv in raw:
        rtrows.append((widx[mon], city, coc, op, rtv))
NW = len(MXW)

# ---- solo cocinas con data real (RTWT turbo internas o ventas); descarta el resto ----
valid_cocs = {x[2] for x in rtrows} | {x[2] for x in vrecs}
stores = [s for s in stores if s['k'] in valid_cocs and s['k'] not in ('—', None)]

CITIES = sorted({s['c'] for s in stores})
OPS = sorted({s['op'] for s in stores if s['op']})

def compute(kind, name):
    incs = lambda s: kind == 'all' or (kind == 'city' and s['c'] == name) or (kind == 'op' and s['op'] == name)
    incv = lambda x: kind == 'all' or (kind == 'city' and x[1] == name) or (kind == 'op' and x[5] == name)
    st = [s for s in stores if incs(s)]
    fins = [s['fin'] for s in st if s['fin'] is not None]
    props = [s['prop'] for s in st if s['prop'] is not None]
    cov = 100*avg(fins)/IDEAL if fins else None
    cov_prop = 100*avg(props)/IDEAL if props else None
    curs = [s['cur'] for s in st if s.get('cur') is not None]
    cov_prev = 100*avg(curs)/IDEAL if curs else None
    dist = Counter(lvl(f) for f in fins); tot = len(fins)
    # ---- RTWT semanal (de KDS time mx) filtrado por scope ----
    rrs = [x for x in rtrows if kind == 'all' or (kind == 'city' and x[1] == name) or (kind == 'op' and x[3] == name)]
    L = NW - 1
    wk_net = []
    for w in range(NW):
        vv = [x[4] for x in rrs if x[0] == w]; wk_net.append(round(avg(vv), 2) if vv else None)
    rtwt = wk_net[L] if NW else None
    rtwt_lw = wk_net[L-1] if NW > 1 else None
    def rt_lp(keyidx, key):
        la = [x[4] for x in rrs if x[0] == L and x[keyidx] == key]
        pr = [x[4] for x in rrs if x[0] == L-1 and x[keyidx] == key]
        return avg(la), avg(pr)
    rt3 = sum(1 for x in rrs if x[0] == L and x[4] > 3)
    alerts = dict(rt3=rt3, poly1=dist.get(1000, 0))

    # distribución de polígonos POR FECHA: semana anterior (Current Size) vs actual (FINAL SIZE)
    d_prev = Counter(lvl(s['cur']) for s in st if s.get('cur') is not None)
    d_curr = Counter(lvl(s['fin']) for s in st if s['fin'] is not None)
    glabels = ([MXW[-2], MXW[-1]] if NW >= 2 else ['Anterior', 'Actual'])
    stack = {str(l): [d_prev.get(l, 0), d_curr.get(l, 0)] for l in LV}
    stack_tot = [sum(d_prev.values()), sum(d_curr.values())]

    def grp_rows(keyf, keyidx):
        by = defaultdict(list)
        for s in st: by[keyf(s)].append(s)
        out = []
        for k, v in by.items():
            f = [x['fin'] for x in v if x['fin'] is not None]
            pp = [x['prop'] for x in v if x['prop'] is not None]
            rr, rl = rt_lp(keyidx, k)     # RTWT última vs anterior (KDS semanal)
            out.append(dict(k=k, city=Counter(x['c'] for x in v).most_common(1)[0][0],
                n=len(v), rtwt=None if rr is None else round(rr, 2),
                drt=None if (rr is None or rl is None) else round(rr-rl, 2),
                cov=None if not f else round(100*avg(f)/IDEAL, 1),
                cov_prop=None if not pp else round(100*avg(pp)/IDEAL, 1),
                dt=[dict(b=x['b'], fin=x['fin'], prop=x['prop']) for x in v]))
        out.sort(key=lambda x: x['cov'] if x['cov'] is not None else 999)
        return out
    cocinas = grp_rows(lambda s: s['k'], 2)   # keyidx 2 = cocina en rtrows
    cities = grp_rows(lambda s: s['c'], 1)    # keyidx 1 = ciudad en rtrows

    # ---- ventas KDS (última vs anterior) ----
    lw = [x for x in lw_all if incv(x)]; pw = [x for x in pw_all if incv(x)]
    # órdenes por ciudad y por cocina (para las tablas)
    ord_city = defaultdict(float); ord_coc = defaultdict(float)
    for x in lw: ord_city[x[1]] += x[4]; ord_coc[x[2]] += x[4]
    for r in cities: r['orders'] = round(ord_city.get(r['k'], 0))
    for r in cocinas: r['orders'] = round(ord_coc.get(r['k'], 0))
    # RTWT/cobertura por cocina (de las tiendas del scope) para el detalle
    coc_rt = {}; coc_cov = {}
    bycoc = defaultdict(list)
    for s in st: bycoc[s['k']].append(s)
    for k, v in bycoc.items():
        coc_rt[k] = rt_lp(2, k)[0]     # RTWT última semana por cocina (KDS)
        f = [x['fin'] for x in v if x['fin'] is not None]
        coc_cov[k] = round(100*avg(f)/IDEAL) if f else None
        coc_fin = {k: (lvl(avg([x['fin'] for x in v if x['fin'] is not None])) if any(x['fin'] is not None for x in v) else None) for k, v in bycoc.items()}
    bt = defaultdict(float); btp = defaultdict(float)
    for x in lw: bt[norm(x[3])] += x[4]
    for x in pw: btp[norm(x[3])] += x[4]
    tot_o = sum(bt.values()) or 1
    pareto = []; cum = 0
    for cb, o in sorted(bt.items(), key=lambda x: -x[1]):
        cum += o
        pareto.append(dict(brand=bdisp.get(cb, cb), cb=cb, orders=round(o),
            pct=round(100*o/tot_o, 1), cum=round(100*cum/tot_o, 1)))
    top5 = [p['cb'] for p in pareto[:5]]
    cbk = defaultdict(float)
    for x in lw: cbk[(x[2], norm(x[3]))] += x[4]
    coc_list = sorted({k[0] for k in cbk})
    per_cocina = {}
    rank_in = {}
    for coc in coc_list:
        items = sorted([(cb, o) for (c, cb), o in cbk.items() if c == coc], key=lambda x: -x[1])
        tt = sum(o for _, o in items) or 1
        per_cocina[coc] = [dict(brand=bdisp.get(cb, cb), orders=round(o), pct=round(100*o/tt, 1)) for cb, o in items[:5]]
        for i, (cb, o) in enumerate(items): rank_in[(coc, cb)] = (i+1, round(o), len(items))
    top5_detail = []
    for cb in top5:
        rows = []
        for coc in coc_list:
            o = cbk.get((coc, cb))
            if not o: continue
            rk, orders, ncb = rank_in[(coc, cb)]
            rt = coc_rt.get(coc); sz = coc_fin.get(coc); cv = coc_cov.get(coc)
            sev = 'bad' if (rt is not None and rt > 3) or (sz is not None and sz <= 1000) else ('warn' if (rt is not None and rt > 2) or (sz is not None and sz <= 2100) else 'ok')
            rows.append(dict(coc=coc, orders=orders, rank=rk, nbrands=ncb,
                rt=None if rt is None else round(rt, 2), size=sz, cov=cv, sev=sev))
        rows.sort(key=lambda x: -x['orders'])
        wr = [r for r in rows if r['rt'] is not None]
        best = min(wr, key=lambda x: x['rt'])['coc'] if wr else None
        worst = max(wr, key=lambda x: x['rt'])['coc'] if wr else None
        ac = [r['cov'] for r in rows if r['cov'] is not None]
        ar = [r['rt'] for r in rows if r['rt'] is not None]
        oprev = btp.get(cb, 0)
        wow = None if not oprev else round(100*(bt[cb]-oprev)/oprev, 1)
        top5_detail.append(dict(brand=bdisp.get(cb, cb), orders=round(bt[cb]), wow=wow,
            avg_cov=round(avg(ac), 1) if ac else None, avg_rt=round(avg(ar), 2) if ar else None,
            best=best, worst=worst, rows=rows))

    return dict(
        kpi=dict(rtwt=None if rtwt is None else round(rtwt, 2),
                 rtwt_lw=None if rtwt_lw is None else round(rtwt_lw, 2),
                 drt=None if (rtwt is None or rtwt_lw is None) else round(rtwt-rtwt_lw, 2),
                 cov=None if cov is None else round(cov, 1),
                 cov_prop=None if cov_prop is None else round(cov_prop, 1),
                 n_stores=len(st), n_orders=round(sum(bt.values())), n_brands=len(pareto),
                 n_cocinas=len({s['k'] for s in st}),
                 dist={str(l): dist.get(l, 0) for l in LV}, tot=tot, alerts=alerts),
        stack=dict(labels=glabels, series=stack, tot=stack_tot),
        weekly_rt=wk_net,
        weekly_cov=([None]*(NW-2) + [None if cov_prev is None else round(cov_prev, 1), None if cov is None else round(cov, 1)]) if NW >= 2 else ([None if cov is None else round(cov, 1)]),
        cities=cities, cocinas=cocinas,
        pareto=pareto, top5=top5, top5_detail=top5_detail,
        per_cocina=per_cocina, coc_list=coc_list)

data = {'ALL': compute('all', None)}
for c in CITIES: data[c] = compute('city', c)
for o in OPS: data['op::'+o] = compute('op', o)

updated = None
for r in ws.iter_rows(min_row=2, values_only=True):
    u = col(r, 'UPDATED_AT_UTC')
    if u: updated = str(u)[:10]; break

MX = dict(updated=updated, ideal=IDEAL, LV=LV, cities=CITIES, ops=OPS, weeks=MXW,
          sales_week=sales_week, prev_week=prevlabel, data=data)
json.dump(MX, open('mx_data.json', 'w'), ensure_ascii=False)
k = data['ALL']['kpi']
print('MX | actualizado', updated, '| ventas', sales_week, 'vs', prevlabel)
print('ciudades', CITIES)
print('ops', OPS)
print('ALL: RTWT', k['rtwt'], 'vs LW', k['rtwt_lw'], '(Δ', k['drt'], ') | cob', k['cov'], '% prop', k['cov_prop'], '% | órdenes', k['n_orders'])

if '--inject' in sys.argv:
    idx = open('index_mx.html', encoding='utf-8').read()
    payload = '/*MX_DATA_START*/const MX=' + json.dumps(MX, ensure_ascii=False) + ';/*MX_DATA_END*/'
    new = re.sub(r'/\*MX_DATA_START\*/.*?/\*MX_DATA_END\*/', lambda m: payload, idx, count=1, flags=re.S)
    assert new != idx
    open('index_mx.html', 'w', encoding='utf-8').write(new)
    print('MX inyectado en index_mx.html')
