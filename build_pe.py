#!/usr/bin/env python3
# Pipeline dashboard PERÚ. Esquema de Colombia pero con la ÚLTIMA SEMANA (no promedio 4 sem).
#  - RTWT: última semana (29-jun) vs anterior (22-jun)   [RTWT Foodology PE Turbo]
#  - Cobertura: polígono asignado por tienda (metros)     [Coberturas Foodology PE Turbo]
#  - Estados por RTWT de la última semana: 🟢<2 🟡2-5 🟠5-8 🔴>8 (pierde cobertura)
#  - coverageResult = cobertura si RTWT última ≤ 8, si no 0 (penalización)
#  - Filtro por zona (San Isidro, Surquillo, La Molina, Gonzáles Prada)
# Genera pe_data.json e inyecta const PE en index_pe.html.
import openpyxl, json, sys, re, unicodedata, datetime, os, warnings
from collections import Counter, defaultdict
warnings.filterwarnings('ignore')

RTWT = 'PE_rtwt.xlsx'
COB = 'PE_cobertura.xlsx'
KDS = 'KDS_ventas_pe.xlsx'
LV = [1000, 2100, 2400, 2700, 3000]
IDEAL = 3000

def num(x): return float(x) if isinstance(x, (int, float)) else None
def to_m(x):
    v = num(x)
    return None if v is None else (v*1000 if v < 50 else v)
def sac(s): return ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
def avg(a):
    a = [x for x in a if x is not None]; return sum(a)/len(a) if a else None
def lvl(m): return min(LV, key=lambda l: abs(l-m)) if m is not None else None
def zona(name):
    n = sac(name).lower()
    if 'san isidro' in n: return 'San Isidro'
    if 'surquillo' in n: return 'Surquillo'
    if 'molina' in n: return 'La Molina'
    if 'gonzalez prada' in n or 'manuel gonzalez' in n or 'gonzales prada' in n: return 'Gonzáles Prada'
    return 'Otra'
def bkey(b):
    n = sac(b).lower(); n = re.sub(r'\bturbo\b', ' ', n).replace('&', ' ').replace('-', ' ')
    t = re.sub(r'[^a-z0-9 ]', ' ', n).split(); return ' '.join(t[:2])
def bdisp(b):
    return re.sub(r'\s*-\s*Turbo.*$', '', str(b or ''), flags=re.I).replace(' Turbo', '').strip()

# ---- RTWT (última=29-jun col5, anterior=22-jun col4) ----
wb = openpyxl.load_workbook(RTWT, read_only=True, data_only=True)
rows = [list(r) for r in wb['Resumen'].iter_rows(values_only=True)]
rt = {}; curbrand = None
for r in rows[4:]:
    if r[0] and str(r[0]).strip() and str(r[0]) != 'Total general': curbrand = r[0]
    if r[1]:
        rt[(bkey(curbrand), zona(r[1]))] = [num(r[2]), num(r[3]), num(r[4]), num(r[5])]  # jun8,15,22,29
WEEKS = ['Jun 8', 'Jun 15', 'Jun 22', 'Jun 29']

# ---- cobertura ----
wc = openpyxl.load_workbook(COB, read_only=True, data_only=True)
stores = {}
for r in list(wc['Base'].iter_rows(values_only=True))[1:]:
    if not r[3]: continue
    z = zona(r[4]); k = (bkey(r[2]), z)
    wk = rt.get(k, [None, None, None, None])
    stores[str(r[3])] = dict(b=bdisp(r[2]), k=z, c='Lima', cc=to_m(r[5]), rtw=wk, rt=wk[3], rtlw=wk[2])

ST = list(stores.values())
ZONAS = sorted({s['k'] for s in ST})

# ---- ventas KDS (WoW) ----
def bnorm(b):
    n = sac(b).lower().replace('&', ' ').replace('-', ' ')
    drop = {'turbo', 'fdl', 'fd', 'court', 'sandwiches'}
    return ' '.join(t for t in re.sub(r'[^a-z0-9 ]', ' ', n).split() if t not in drop).strip()
def bshow(b):
    s = re.sub(r'\s*-?\s*turbo\b.*$', '', str(b or ''), flags=re.I).strip()
    return (s.title() if s.isupper() else s) or str(b)
vrecs = []
if os.path.exists(KDS):
    wv = openpyxl.load_workbook(KDS, read_only=True, data_only=True)
    ws = wv['Export']
    for r in ws.iter_rows(min_row=2, values_only=True):
        r = list(r) + [None]*(6-len(list(r)))
        if not isinstance(r[0], datetime.datetime): continue
        vrecs.append((r[0].date(), zona(r[2]), str(r[3] or ''), num(r[4]) or 0))
days = [x[0] for x in vrecs]
sales_week = prev_week = None; lw_all = pw_all = []
vdisp = {}
if days:
    lastday = max(days); firstlw = lastday - datetime.timedelta(days=6)
    pfirst = firstlw - datetime.timedelta(days=7); plast = firstlw - datetime.timedelta(days=1)
    sales_week = f"{firstlw.strftime('%d/%m')}–{lastday.strftime('%d/%m/%Y')}"
    prev_week = f"{pfirst.strftime('%d/%m')}–{plast.strftime('%d/%m')}"
    lw_all = [x for x in vrecs if firstlw <= x[0] <= lastday]
    pw_all = [x for x in vrecs if pfirst <= x[0] <= plast]
    for x in lw_all + pw_all: vdisp.setdefault(bnorm(x[2]), bshow(x[2]))

def covResult(s):
    if s['cc'] is None: return None
    return 0 if (s['rt'] is not None and s['rt'] > 8) else s['cc']

# índice tienda por (primer token de marca, zona) para el detalle de ventas
def ftok(b):
    t = bnorm(b).split(); return t[0] if t else ''
sidx = {}
for s in ST:
    sidx[(ftok(s['b']), s['k'])] = (s['rt'], s['cc'], covResult(s))

def compute(kind, name):
    st = [s for s in ST if kind == 'all' or s['k'] == name]
    fins = [s['cc'] for s in st if s['cc'] is not None]
    rts = [s['rt'] for s in st if s['rt'] is not None]
    rtwt = avg(rts); rtwt_lw = avg([s['rtlw'] for s in st if s['rtlw'] is not None])
    crs = [covResult(s) for s in st if covResult(s) is not None]
    cov = 100*avg(crs)/IDEAL if crs else None
    cov_asg = 100*avg(fins)/IDEAL if fins else None
    dist = Counter(lvl(f) for f in fins); tot = len(fins)
    lose = sum(1 for s in st if s['rt'] is not None and s['rt'] > 8)
    alerts = dict(rt3=sum(1 for x in rts if x > 3), poly1=dist.get(1000, 0))
    weekly = [ (lambda a: round(a, 2) if a is not None else None)(avg([s['rtw'][w] for s in st if s.get('rtw')])) for w in range(4) ]
    # distribución apilada por zona
    gg = defaultdict(list)
    for s in st:
        if s['cc'] is not None: gg[s['k']].append(s)
    glabels = sorted(gg.keys(), key=lambda k: 100*avg([x['cc'] for x in gg[k]])/IDEAL)
    stack = {str(l): [] for l in LV}; stack_tot = []
    for lab in glabels:
        d = Counter(lvl(x['cc']) for x in gg[lab])
        for l in LV: stack[str(l)].append(d.get(l, 0))
        stack_tot.append(sum(d.values()))
    # tabla por zona
    by = defaultdict(list)
    for s in st: by[s['k']].append(s)
    cocinas = []
    for k, v in by.items():
        rr = avg([x['rt'] for x in v if x['rt'] is not None]); rl = avg([x['rtlw'] for x in v if x['rtlw'] is not None])
        cr = [covResult(x) for x in v if covResult(x) is not None]
        cocinas.append(dict(k=k, city='Lima', n=len(v),
            rtwt=None if rr is None else round(rr, 2),
            drt=None if (rr is None or rl is None) else round(rr-rl, 2),
            cov=None if not cr else round(100*avg(cr)/IDEAL, 1),
            lose=sum(1 for x in v if x['rt'] is not None and x['rt'] > 8)))
    cocinas.sort(key=lambda x: x['cov'] if x['cov'] is not None else 999)
    # tiendas fuera de cobertura (rt última > 8)
    lc = sorted([dict(b=s['b'], k=s['k'], rt=s['rt'], cc=s['cc'], res=covResult(s)) for s in st if s['rt'] is not None and s['rt'] > 8],
                key=lambda x: -x['rt'])
    # RTWT en aumento (2+ semanas seguidas subiendo, hasta la última: jun 29 = índice 3)
    rising = []
    L = 3
    for s in st:
        w = s.get('rtw') or []
        if len(w) < 4 or w[L] is None or w[L-1] is None or not (w[L] > w[L-1]): continue
        i = L; path = [w[L]]; inc = 0
        while i-1 >= 0 and w[i-1] is not None and w[i] > w[i-1]:
            path.insert(0, round(w[i-1], 2)); inc += 1; i -= 1
        if inc >= 2: rising.append(dict(b=s['b'], k=s['k'], inc=inc, path=[round(x, 2) for x in path], last=round(w[L], 2)))
    rising.sort(key=lambda x: -x['last'])
    # tiendas en polígono mínimo (cobertura efectiva = 1.0 km)
    min1 = sorted([dict(b=s['b'], k=s['k'], rt=None if s['rt'] is None else round(s['rt'], 2)) for s in st if covResult(s) == 1000],
                  key=lambda x: -(x['rt'] or 0))
    # marcas con peor RTWT de la última semana
    byb = defaultdict(list)
    for s in st:
        if s['rt'] is not None: byb[s['b']].append(s['rt'])
    brand_worst = sorted([dict(brand=b, rt=round(avg(v), 2), n=len(v)) for b, v in byb.items()], key=lambda x: -x['rt'])[:15]
    # ---- ventas KDS (última vs anterior) ----
    vinc = lambda x: kind == 'all' or (kind == 'zona' and x[1] == name)
    lw = [x for x in lw_all if vinc(x)]; pw = [x for x in pw_all if vinc(x)]
    bt = defaultdict(float); btp = defaultdict(float)
    for x in lw: bt[bnorm(x[2])] += x[3]
    for x in pw: btp[bnorm(x[2])] += x[3]
    tot_o = sum(bt.values()) or 1
    pareto = []; cum = 0
    for cb, o in sorted(bt.items(), key=lambda x: -x[1]):
        cum += o; pareto.append(dict(brand=vdisp.get(cb, cb), orders=round(o), pct=round(100*o/tot_o, 1), cum=round(100*cum/tot_o, 1)))
    top5c = [p for p in sorted(bt, key=lambda c: -bt[c])[:5]]
    cbz = defaultdict(float)
    for x in lw: cbz[(x[1], bnorm(x[2]))] += x[3]
    zonas_v = sorted({k[0] for k in cbz})
    per_zona = {}; rankz = {}
    for z in zonas_v:
        items = sorted([(cb, o) for (zz, cb), o in cbz.items() if zz == z], key=lambda x: -x[1])
        tt = sum(o for _, o in items) or 1
        per_zona[z] = [dict(brand=vdisp.get(cb, cb), orders=round(o), pct=round(100*o/tt, 1)) for cb, o in items[:5]]
        for i, (cb, o) in enumerate(items): rankz[(z, cb)] = (i+1, round(o), len(items))
    top5_detail = []
    for cb in top5c:
        rr = []
        for z in zonas_v:
            o = cbz.get((z, cb))
            if not o: continue
            rk, orders, nb = rankz[(z, cb)]
            rtv, cc, cr = sidx.get((cb.split()[0] if cb else '', z), (None, None, None))   # RTWT/polígono por MARCA+zona
            cv = None if cr is None else round(100*cr/IDEAL)
            sz = lvl(cr) if cr else None
            sev = 'bad' if (rtv is not None and rtv > 3) or (sz is not None and sz <= 1000) else ('warn' if (rtv is not None and rtv > 2) or (sz is not None and sz <= 2100) else 'ok')
            rr.append(dict(coc=z, orders=orders, rank=rk, nbrands=nb, rt=None if rtv is None else round(rtv, 2), size=sz, cov=cv, sev=sev))
        rr.sort(key=lambda x: -x['orders'])
        wr = [r for r in rr if r['rt'] is not None]
        best = min(wr, key=lambda x: x['rt'])['coc'] if wr else None
        worst = max(wr, key=lambda x: x['rt'])['coc'] if wr else None
        ar = [r['rt'] for r in rr if r['rt'] is not None]; ac = [r['cov'] for r in rr if r['cov'] is not None]
        oprev = btp.get(cb, 0); wow = None if not oprev else round(100*(bt[cb]-oprev)/oprev, 1)
        top5_detail.append(dict(brand=vdisp.get(cb, cb), orders=round(bt[cb]), wow=wow,
            avg_rt=round(avg(ar), 2) if ar else None, avg_cov=round(avg(ac), 1) if ac else None,
            best=best, worst=worst, rows=rr))
    top5_share = round(100*sum(bt[c] for c in top5c)/tot_o, 1) if bt else 0
    n80 = next((i+1 for i, p in enumerate(pareto) if p['cum'] >= 80), len(pareto))
    ventas = dict(pareto=pareto, top5=[vdisp.get(c, c) for c in top5c], top5_detail=top5_detail,
                  per_zona=per_zona, zonas_v=zonas_v, total_orders=round(sum(bt.values())),
                  n_brands=len(pareto), top5_share=top5_share, n80=n80)
    return dict(
        kpi=dict(rtwt=None if rtwt is None else round(rtwt, 2),
                 rtwt_lw=None if rtwt_lw is None else round(rtwt_lw, 2),
                 drt=None if (rtwt is None or rtwt_lw is None) else round(rtwt-rtwt_lw, 2),
                 cov=None if cov is None else round(cov, 1),
                 cov_asg=None if cov_asg is None else round(cov_asg, 1),
                 n_stores=len(st), n_cocinas=len({s['k'] for s in st}),
                 dist={str(l): dist.get(l, 0) for l in LV}, tot=tot, lose=lose, alerts=alerts),
        stack=dict(labels=glabels, series=stack, tot=stack_tot),
        weekly=weekly,
        cocinas=cocinas, lost=lc, rising=rising, min1=min1, brand_worst=brand_worst, ventas=ventas)

data = {'ALL': compute('all', None)}
for z in ZONAS: data[z] = compute('zona', z)
PE = dict(ideal=IDEAL, LV=LV, zonas=ZONAS, weeks=WEEKS, week='29-jun', prev_week='22-jun',
          sales_week=sales_week, sales_prev=prev_week, data=data)
json.dump(PE, open('pe_data.json', 'w'), ensure_ascii=False)
k = data['ALL']['kpi']
print('PE | zonas', ZONAS, '| tiendas', len(ST))
print('ALL: RTWT última', k['rtwt'], 'vs prev', k['rtwt_lw'], '(Δ', k['drt'], ') | cob efec', k['cov'], '% asignada', k['cov_asg'], '% | pierden cob', k['lose'])
print('dist', k['dist'])

if '--inject' in sys.argv:
    idx = open('index_pe.html', encoding='utf-8').read()
    payload = '/*PE_DATA_START*/const PE=' + json.dumps(PE, ensure_ascii=False) + ';/*PE_DATA_END*/'
    new = re.sub(r'/\*PE_DATA_START\*/.*?/\*PE_DATA_END\*/', lambda m: payload, idx, count=1, flags=re.S)
    assert '/*PE_DATA_START*/' in new, 'marcadores PE no encontrados'
    open('index_pe.html', 'w', encoding='utf-8').write(new)
    print('PE inyectado en index_pe.html' + ('' if new != idx else ' (sin cambios en la data)'))
