#!/usr/bin/env python3
"""
Regenera el bloque de datos de VENTAS embebido en index.html a partir del
Google Sheet "Turbology" (hojas 'Raw' y 'Raw ventas').

Uso:
  python3 build_ventas.py Turbology.xlsx   # produce ventas_data.json e inyecta en index.html

La marca de datos (fecha de ventas usada, etc.) queda en el JSON.
"""
import sys, json, csv, re, unicodedata, datetime, os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "KDS_ventas.xlsx")
XLSX_LIST = sys.argv[1:] if len(sys.argv) > 1 else [XLSX]   # uno o varios KDS (se concatenan)
STORE2COC = os.path.join(HERE, "store2cocina.json")
COC2OP = os.path.join(HERE, "cocina2ops.json")

# ---- normalización de marcas (une ventas <-> RTWT) ----
def norm(s):
    if s is None: return ""
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii','ignore').decode().lower()
    s = s.replace('&',' ').replace('.',' ').replace('-',' ')
    s = re.sub(r'\bby\b.*$', ' ', s)          # "... by maluma/dorito/robegrill"
    toks_drop = {'turbo','fdl','fd','court','sandwiches','pase'}
    s = ' '.join(t for t in re.sub(r'[^a-z0-9 ]',' ',s).split() if t not in toks_drop)
    return s.strip()

ALIAS = {  # norm ventas -> norm RTWT canónico
    'el jefe':'sanduches el jefe',
    'san jeronimo':'helados san jeronimo',
    'almuerzos de la casa':'almuerzos d la casa',
    'cinnabon foodology':'cinnabon',
}
def canon(s):
    n = norm(s); return ALIAS.get(n, n)

# ---- cocina desde kitchen_id de ventas ("69 ARRECIFE" -> "Arrecife") ----
def coc_norm(s):
    s = unicodedata.normalize('NFKD', str(s or '')).encode('ascii','ignore').decode().lower()
    return re.sub(r'[^a-z0-9 ]',' ', s).strip()
def kitchen_to_cocina(kid, cocina_names):
    base = re.sub(r'^\s*\d+\s*', '', str(kid)).strip()       # quita número
    base = base.replace(' CINNABON','')
    nb = coc_norm(base)
    for c in cocina_names:
        cn = coc_norm(c)
        if cn == nb or nb in cn or cn.split(' /')[0]==nb: return c
    return base.title()

def to_float(x):
    try: return float(str(x).replace(',',''))
    except: return None

def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    store2coc = json.load(open(STORE2COC, encoding='utf-8')) if os.path.exists(STORE2COC) else {}
    coc2op = json.load(open(COC2OP, encoding='utf-8')) if os.path.exists(COC2OP) else {}

    # ---------- RTWT / cobertura: data ACTUAL (const D = D_v2.json) ----------
    # Usa la misma data que el resto del dashboard: RTWT de la última semana vs la anterior,
    # y el polígono asignado (cc). Así las cards de ventas quedan consistentes con la cobertura real.
    Dj = json.load(open(os.path.join(HERE, 'D_v2.json'), encoding='utf-8'))
    DW = Dj['weeks']; LIw = len(DW) - 1
    rt_rows=[]; rt_prev_rows=[]
    for s in Dj['stores']:
        cb=canon(s.get('b')); coc=s.get('k') or 'Otra'; city=s.get('c'); op=s.get('op'); cc=s.get('cc')
        _rt=s.get('rt', [])
        rt_cur = _rt[LIw]   if LIw   < len(_rt) else None
        rt_prv = _rt[LIw-1] if LIw-1 < len(_rt) else None
        rt_rows.append(dict(cb=cb, coc=coc, city=city, rt=rt_cur, fin=cc, op=op))
        rt_prev_rows.append(dict(cb=cb, coc=coc, city=city, rt=rt_prv, fin=cc, op=op))
    rtwt_week = DW[LIw]

    # ---------- VENTAS (uno o varios KDS: hoja Export) ----------
    recs=[]; days=set(); has_date=False
    for _fp in XLSX_LIST:
        _wb = wb if _fp == XLSX else openpyxl.load_workbook(_fp, read_only=True, data_only=True)
        _sh = 'Export' if 'Export' in _wb.sheetnames else ('Raw ventas' if 'Raw ventas' in _wb.sheetnames else _wb.sheetnames[0])
        _v = list(_wb[_sh].iter_rows(values_only=True))
        _vi = {str(h).strip().lower():i for i,h in enumerate(_v[0]) if h}
        def _c(row,name,_vi=_vi):
            i=_vi.get(name.lower()); return row[i] if (i is not None and i<len(row)) else None
        _datecol = 'date' if 'date' in _vi else ('week_year' if 'week_year' in _vi else None)
        _shift = 6 if _datecol == 'week_year' else 0     # week_year = lunes -> domingo (nuestra convención)
        if _datecol: has_date=True
        for r in _v[1:]:
            brand=_c(r,'brand') or _c(r,'brand_name')
            if brand in (None, ''): continue
            dd=None
            if _datecol:
                d=_c(r,_datecol)
                if not isinstance(d, datetime.datetime):
                    try: d=datetime.datetime.fromisoformat(str(d))
                    except: continue
                dd=(d+datetime.timedelta(days=_shift)).date(); days.add(dd)
            kit=_c(r,'kitchen_id') or _c(r,'kitchen')     # nombre/id de cocina (según export)
            orders=to_float(_c(r,'Total orders')) or to_float(_c(r,'orders')) or 0
            recs.append((dd, _c(r,'city'), kit, brand, orders, _c(r,'ops')))
    if has_date and days:
        # totales por semana; descarta la última semana si está incompleta (< 50% de la anterior)
        _wtot={}
        for _x in recs:
            if _x[0] is not None: _wtot[_x[0]]=_wtot.get(_x[0],0)+_x[4]
        _sd=sorted(_wtot)
        while len(_sd)>=2 and _wtot[_sd[-1]] < 0.5*_wtot[_sd[-2]]: _sd.pop()
        lastday=_sd[-1]; firstlw=lastday-datetime.timedelta(days=6)
        sales_week=f"{firstlw.strftime('%d/%m')}–{lastday.strftime('%d/%m/%Y')}"
        pfirst=firstlw-datetime.timedelta(days=7); plast=firstlw-datetime.timedelta(days=1)
        lw_all=[x for x in recs if x[0] is not None and firstlw<=x[0]<=lastday]
        pw_all=[x for x in recs if x[0] is not None and pfirst<=x[0]<=plast]
    else:
        # archivo sin fechas: totales actuales, sin comparativo WoW
        sales_week='acumulado (KDS 30 jul)'
        lw_all=recs; pw_all=[]

    cocina_names=sorted(set(x['coc'] for x in rt_rows))
    def pretty(name): return name.title() if name.isupper() else name
    disp={}
    for _,_,_,b,_,_ in lw_all: disp.setdefault(canon(b), pretty(b.strip()))

    def indexrt(rows):
        bc={}
        for x in rows:
            d=bc.setdefault((x['cb'],x['coc']), {'rt':[], 'fin':[]})
            if x['rt'] is not None: d['rt'].append(x['rt'])
            if x['fin'] is not None: d['fin'].append(x['fin'])
        return bc
    def aggf(bc,key):
        d=bc.get(key)
        if not d: return (None,None)
        rt=sum(d['rt'])/len(d['rt']) if d['rt'] else None
        fin=sum(d['fin'])/len(d['fin']) if d['fin'] else None
        return (rt,fin)

    def compute(scope):
        kind,name=scope
        incv=lambda x: (kind=='all') or (kind=='city' and x[1]==name) or (kind=='op' and x[5]==name)
        incr=lambda r: (kind=='all') or (kind=='city' and r['city']==name) or (kind=='op' and r.get('op')==name)
        lw=[x for x in lw_all if incv(x)]
        if not lw: return None
        pw=[x for x in pw_all if incv(x)]
        bc=indexrt([r for r in rt_rows if incr(r)])
        bcp=indexrt([r for r in rt_prev_rows if incr(r)])
        agg=lambda k: aggf(bc,k); agg_prev=lambda k: aggf(bcp,k)
        bt={}
        for _,_,_,b,o,_ in lw: bt[canon(b)]=bt.get(canon(b),0)+o
        tot=sum(bt.values()) or 1
        bt_prev={}
        for _,_,_,b,o,_ in pw: bt_prev[canon(b)]=bt_prev.get(canon(b),0)+o
        pareto=[]; cum=0
        for cb,o in sorted(bt.items(), key=lambda x:-x[1]):
            cum+=o
            pareto.append(dict(brand=disp.get(cb,cb), cb=cb, orders=round(o),
                               pct=round(100*o/tot,1), cum=round(100*cum/tot,1)))
        top5=[p['cb'] for p in pareto[:5]]
        cbk={}
        for _,_,kid,b,o,_ in lw:
            coc=kitchen_to_cocina(kid, cocina_names)
            cbk[(coc,canon(b))]=cbk.get((coc,canon(b)),0)+o
        cocinas=sorted(set(k[0] for k in cbk))
        per_cocina={}; rank_in={}
        for coc in cocinas:
            items=sorted([(cb,o) for (c,cb),o in cbk.items() if c==coc], key=lambda x:-x[1])
            tt=sum(o for _,o in items) or 1
            per_cocina[coc]=[]
            for cb,o in items[:5]:
                rtv,finv=agg((cb,coc))
                per_cocina[coc].append(dict(brand=disp.get(cb,cb), cb=cb, orders=round(o), pct=round(100*o/tt,1),
                                            rt=None if rtv is None else round(rtv,2),
                                            size=None if finv is None else round(finv)))
            for i,(cb,o) in enumerate(items): rank_in[(coc,cb)]=(i+1,round(o),len(items))
        top5_detail=[]
        for cb in top5:
            rows=[]
            for coc in cocinas:
                o=cbk.get((coc,cb))
                if not o: continue
                rk,orders,ncb=rank_in[(coc,cb)]
                rt,fin=agg((cb,coc))
                covpct=None if fin is None else round(100*fin/3000,1)
                sev='ok'
                if (rt is not None and rt>3) or (fin is not None and fin<=1000): sev='bad'
                elif (rt is not None and rt>2) or (fin is not None and fin<=2100): sev='warn'
                rows.append(dict(coc=coc, orders=orders, rank=rk, nbrands=ncb,
                                 rt=None if rt is None else round(rt,2),
                                 size=None if fin is None else round(fin),
                                 cov=covpct, sev=sev))
            rows.sort(key=lambda x:-x['orders'])
            wr=[r for r in rows if r['rt'] is not None]
            best=min(wr,key=lambda x:x['rt'])['coc'] if wr else None
            worst=max(wr,key=lambda x:x['rt'])['coc'] if wr else None
            ar=[r['rt'] for r in rows if r['rt'] is not None]
            ac=[r['cov'] for r in rows if r['cov'] is not None]
            avg_rt=round(sum(ar)/len(ar),2) if ar else None
            avg_cov=round(sum(ac)/len(ac),1) if ac else None
            pr=[]; pc=[]
            for r in rows:
                rtp,finp=agg_prev((cb,r['coc']))
                if rtp is not None: pr.append(rtp)
                if finp is not None: pc.append(100*finp/3000)
            avg_rt_prev=sum(pr)/len(pr) if pr else None
            avg_cov_prev=sum(pc)/len(pc) if pc else None
            drt=None if (avg_rt is None or avg_rt_prev is None) else round(avg_rt-avg_rt_prev,2)
            dcov=None if (avg_cov is None or avg_cov_prev is None) else round(avg_cov-avg_cov_prev,1)
            oprev=bt_prev.get(cb,0)
            wow=None if not oprev else round(100*(bt[cb]-oprev)/oprev,1)
            top5_detail.append(dict(brand=disp.get(cb,cb), cb=cb, orders=round(bt[cb]),
                wow=wow, orders_prev=round(oprev), avg_rt=avg_rt, avg_cov=avg_cov,
                drt=drt, dcov=dcov, best=best, worst=worst, rows=rows))
        return dict(meta=dict(total_orders=round(tot), n_brands=len(pareto),
                    top5_share=round(sum(bt[c] for c in top5)/tot*100,1)),
                    pareto=pareto, top5=top5, top5_detail=top5_detail,
                    per_cocina=per_cocina, cocinas=cocinas)

    cities=sorted(set(x[1] for x in lw_all if x[1]))
    ops=sorted(set(x[5] for x in lw_all if x[5]))
    data={'ALL':compute(('all',None))}
    for c in cities:
        rc=compute(('city',c))
        if rc: data[c]=rc
    for o in ops:
        rc=compute(('op',o))
        if rc: data['op::'+o]=rc
    payload=dict(sales_week=sales_week, rtwt_week=rtwt_week,
                 generated=datetime.date.today().isoformat(),
                 cities=['ALL']+cities, ops=ops, data=data)

    json.dump(payload, open(os.path.join(HERE,'ventas_data.json'),'w'), ensure_ascii=False, indent=1)

    # ---- inyecta el bloque de datos en index.html (entre marcadores) ----
    idx=os.path.join(HERE,'index_co.html')
    if os.path.exists(idx):
        html=open(idx,encoding='utf-8').read()
        A='/*VENTAS_DATA_START*/'; B='/*VENTAS_DATA_END*/'
        block=A+'const VENTAS='+json.dumps(payload,ensure_ascii=False)+';'+B
        if A in html and B in html:
            html=re.sub(re.escape(A)+'.*?'+re.escape(B), lambda m: block, html, flags=re.S)
            open(idx,'w',encoding='utf-8').write(html)
            print('index.html: bloque VENTAS actualizado')
        else:
            print('index.html: marcadores no encontrados (aún no inyectado el HTML de Ventas)')

    # resumen validación
    A=payload['data']['ALL']
    print("VENTAS", sales_week, "| RTWT", rtwt_week, "| ciudades:", payload['cities'])
    for ck in payload['cities']:
        D=payload['data'].get(ck)
        if not D: continue
        n80=next((i+1 for i,p in enumerate(D['pareto']) if p['cum']>=80), len(D['pareto']))
        print(f"  [{ck:12}] orders={D['meta']['total_orders']:5} marcas={D['meta']['n_brands']:2} top5={D['meta']['top5_share']:5.1f}% n80={n80} top5={[t['brand'] for t in D['top5_detail']]}")
    return payload

if __name__=='__main__':
    main()
