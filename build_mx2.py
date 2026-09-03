#!/usr/bin/env python3
# Pipeline dashboard MÉXICO v2.
#  - RTWT: histórico 8 semanas (pestaña 'Trazabilidad RTWT' del reporte del 24 ago).
#  - Etiquetas de semana con +6 (domingo de cierre), como Colombia.
#  - Polígono por semana: Aug9=Size Proposal(3ago) · Aug16=Current Size(17ago) ·
#    Aug23=FINAL SIZE(17ago) · Aug30=FINAL SIZE(24ago). En metros.
#  - Ventas: warehouse (performance_stores, MEX·Rappi·Turbo), última semana Aug24 (24-30) vs Aug17.
# Genera mx_data.json e inyecta const MX en index_mx.html.
import json, re, unicodedata, os, csv, sys
from collections import Counter, defaultdict
import openpyxl, warnings
warnings.filterwarnings('ignore')

LV = [1000, 2100, 2400, 2700, 3000]
IDEAL = 3000

def num(x):
    try: return float(x)
    except: return None
def to_m(x):
    v = num(x)
    return None if v is None else (v*1000 if v < 50 else v)
def sac(s): return ''.join(c for c in unicodedata.normalize('NFD', str(s or '')) if unicodedata.category(c) != 'Mn')
def avg(a):
    a = [x for x in a if x is not None]; return sum(a)/len(a) if a else None
def lvl(m): return min(LV, key=lambda l: abs(l-m)) if m is not None else None
CITY_MAP = {'Ciudad de México': 'CDMX', 'Merida': 'Mérida'}
def ncity(c): return CITY_MAP.get(str(c), str(c))
def stripbrand(s): return re.sub(r'\s*-\s*Turbo\s*$', '', str(s or ''), flags=re.I).strip()
def norm(s):
    if s is None: return ''
    s = sac(s).lower().replace('&', ' ').replace('.', ' ').replace('-', ' ')
    drop = {'turbo', 'fdl', 'fd', 'court', 'sandwiches', 'pase', 'desayunos', 'desayuno'}
    return ' '.join(t for t in re.sub(r'[^a-z0-9 ]', ' ', s).split() if t not in drop).strip()

s2c = json.load(open('store2cocina_mx.json'))
coc2op = json.load(open('cocina2ops_mx.json')) if os.path.exists('cocina2ops_mx.json') else {}
def coc_of(sid): return s2c.get('MX'+sid) or s2c.get(sid) or s2c.get(sid+'.0') or '—'
COCS = set(s2c.values())
def kit2coc(kid):
    nb = sac(re.sub(r'^\s*\d+\s*', '', str(kid))).lower().strip()
    exact = [c for c in COCS if sac(c).lower() == nb]          # 1) match exacto
    if exact: return exact[0]
    for c in COCS:                                             # 2) fallback substring
        cc = sac(c).lower()
        if nb in cc or cc.split(' /')[0] == nb: return c
    return re.sub(r'^\s*\d+\s*', '', str(kid)).title()

# ---------- RTWT histórico + polígonos ----------
RT = json.load(open('mx_rtwt_hist.json'))
POLY = json.load(open('mx_poly.json'))
MXW = RT['labels']              # 8 etiquetas domingo (+6): Jul 12 ... Aug 30
NW = len(MXW)
L = NW - 1

# --- polígono viejo de julio (FOODOLOGY.xlsx): Current Size -> Jul 19, FINAL SIZE -> Jul 26 ---
def _sidn(s):
    if s is None: return None
    s = re.sub(r'\.0+$', '', str(s).strip()); return re.sub(r'\D', '', s) or None
OLDPOLY = {}
if os.path.exists('FOODOLOGY.xlsx'):
    _wb = openpyxl.load_workbook('FOODOLOGY.xlsx', data_only=True); _ws = _wb['DETALLE']
    _h = {c: i for i, c in enumerate(next(_ws.iter_rows(min_row=1, max_row=1, values_only=True)))}
    for _r in _ws.iter_rows(min_row=2, values_only=True):
        _sid = _sidn(_r[_h['Store ID']]) if 'Store ID' in _h else None
        if not _sid: continue
        _cur = _r[_h['Current Size']] if 'Current Size' in _h else None
        _fin = _r[_h['FINAL SIZE']] if 'FINAL SIZE' in _h else None
        OLDPOLY[_sid] = (to_m(_cur), to_m(_fin))

stores = []
rtrows = []                     # (wi, city, cocina, op, rtwt)
for sid, info in RT['stores'].items():
    coc = coc_of(sid)
    p = POLY.get(sid, {})
    city = ncity(p.get('city') or '—')
    op = coc2op.get(coc)
    rtarr = info['rt']          # 8 valores
    pw = [None, None, None, None, to_m(p.get('a9')), to_m(p.get('a16')), to_m(p.get('a23')), to_m(p.get('a30'))]
    # dos semanas de julio con polígono real (FOODOLOGY): Jul 19 = Current Size, Jul 26 = FINAL SIZE
    _op = OLDPOLY.get(sid)
    if _op:
        if _op[0] is not None: pw[1] = _op[0]   # Jul 19
        if _op[1] is not None: pw[2] = _op[1]   # Jul 26
    # semanas de julio sin polígono -> se deja el último dato conocido (arrastre):
    #   forward-fill huecos internos + back-fill el inicio con el primer dato disponible (Aug 9)
    _last = None
    for _i in range(NW):
        if pw[_i] is not None: _last = pw[_i]
        elif _last is not None: pw[_i] = _last
    _first = next((v for v in pw if v is not None), None)
    pw = [v if v is not None else _first for v in pw]
    _pvals = [v for v in pw if v is not None]
    mxp = max(_pvals) if _pvals else None       # polígono máximo de la zona (mayor de las 8 semanas)
    stores.append(dict(sid=sid, b=stripbrand(info.get('marca')), k=coc, c=city, op=op,
        rt=rtarr[L], rtlw=rtarr[L-1], cur=pw[L-1], fin=pw[L], prop=None, rtw=rtarr, pw=pw, mx=mxp))
    for w, v in enumerate(rtarr):
        if v is not None: rtrows.append((w, city, coc, op, v))

CITIES = sorted({s['c'] for s in stores})
OPS = sorted({s['op'] for s in stores if s['op']})
coc2city = {}
for s in stores: coc2city.setdefault(s['k'], s['c'])

# ---------- ventas (warehouse) ----------
CUR, PREV = '2026-08-24', '2026-08-17'
sales_week = '24/08–30/08/2026'
prev_week = '17/08–23/08'
coc_tot = {CUR: {}, PREV: {}}          # cocina -> orders
bra_tot = {CUR: {}, PREV: {}}          # brand(norm) -> orders ; guardamos display
bdisp = {}
if os.path.exists('mx_ventas_agg.csv'):
    for row in csv.DictReader(open('mx_ventas_agg.csv')):
        wk = '2026-08-24' if row['week'] == CUR else ('2026-08-17' if row['week'] == PREV else row['week'])
        wk = row['week']; o = num(row['orders']) or 0
        if row['typ'] == 'COC':
            coc = kit2coc(row['name']); coc_tot.setdefault(wk, {}); coc_tot[wk][coc] = coc_tot[wk].get(coc, 0)+o
        elif row['typ'] == 'BRA':
            cb = norm(row['name']); bdisp.setdefault(cb, stripbrand(row['name']))
            bra_tot.setdefault(wk, {}); bra_tot[wk][cb] = bra_tot[wk].get(cb, 0)+o
# top marcas por cocina (última semana)
percoc = defaultdict(list)             # cocina -> [(brand_norm, orders)]
if os.path.exists('mx_ventas_percoc.csv'):
    for row in csv.DictReader(open('mx_ventas_percoc.csv')):
        coc = kit2coc(row['kitchen']); cb = norm(row['brand'])
        bdisp.setdefault(cb, stripbrand(row['brand']))
        percoc[coc].append((cb, num(row['orders']) or 0))

# ---------- GMV semanal por cocina (warehouse) para $ perdido ----------
MONS = RT['mondays']                    # 8 lunes alineados a MXW
gmvw = defaultdict(lambda: [None]*NW)   # cocina -> [gmv por semana]
if os.path.exists('mx_gmv.csv'):
    for row in csv.DictReader(open('mx_gmv.csv')):
        coc = kit2coc(row['kitchen'])
        if row['week'] in MONS: gmvw[coc][MONS.index(row['week'])] = num(row['gmv'])

# ---------- agregador por scope ----------
def compute(kind, name):
    incs = lambda s: kind == 'all' or (kind == 'city' and s['c'] == name) or (kind == 'op' and s['op'] == name)
    st = [s for s in stores if incs(s)]
    if not st: return None
    cocset = {s['k'] for s in st}
    # --- polígono (Aug30 = fin, Aug23 = cur) ---
    fins = [s['fin'] for s in st if s['fin'] is not None]
    curs = [s['cur'] for s in st if s['cur'] is not None]
    cov = 100*avg([s['fin']/s['mx'] for s in st if s['fin'] is not None and s.get('mx')]) if fins else None
    cov_prev = 100*avg([s['cur']/s['mx'] for s in st if s['cur'] is not None and s.get('mx')]) if curs else None
    dist = Counter(lvl(f) for f in fins)
    d_prev = Counter(lvl(c) for c in curs)
    stack = {str(l): [d_prev.get(l, 0), dist.get(l, 0)] for l in LV}
    stack_tot = [sum(d_prev.values()), sum(dist.values())]
    # --- RTWT semanal (8) ---
    rrs = [x for x in rtrows if kind == 'all' or (kind == 'city' and x[1] == name) or (kind == 'op' and x[3] == name)]
    wk_net = []
    for w in range(NW):
        vv = [x[4] for x in rrs if x[0] == w]; wk_net.append(round(avg(vv), 2) if vv else None)
    rtwt = wk_net[L]; rtwt_lw = wk_net[L-1]
    def rt_lp(keyidx, key):
        la = [x[4] for x in rrs if x[0] == L and x[keyidx] == key]
        pr = [x[4] for x in rrs if x[0] == L-1 and x[keyidx] == key]
        return avg(la), avg(pr)
    rt3 = sum(1 for x in rrs if x[0] == L and x[4] > 3)
    alerts = dict(rt3=rt3, poly1=dist.get(1000, 0))
    # --- cobertura semanal (4 semanas con polígono: idx 4..7) ---
    weekly_cov = []
    for w in range(NW):
        pv = [s['pw'][w]/s['mx'] for s in st if s['pw'][w] is not None and s.get('mx')]
        weekly_cov.append(round(100*avg(pv), 1) if pv else None)
    # --- $ perdido (modelo lineal vs polígono máximo, a nivel cocina) ---
    bycoc_all = defaultdict(list)
    for s in st: bycoc_all[s['k']].append(s)
    def coc_lost(v, w):
        g = gmvw.get(v[0]['k'], [None]*NW)[w] if v else None
        if g is None: return None, None, None
        polys = [x['pw'][w] for x in v if x['pw'][w] is not None]
        maxs = [x['mx'] for x in v if x.get('mx')]
        if not polys or not maxs: return g, None, None
        ap, mp = avg(polys), avg(maxs)
        lo = g*(mp/ap - 1) if (ap and mp and ap < mp) else 0
        return g, (ap/mp if mp else None), lo
    weekly_lost = []
    for w in range(NW):
        tw = 0.0
        for c, v in bycoc_all.items():
            _g, _r, _lo = coc_lost(v, w)
            if _lo: tw += _lo
        weekly_lost.append(round(tw))
    lost = weekly_lost[L]
    gmv_now = round(sum((gmvw.get(c, [None]*NW)[L] or 0) for c in bycoc_all))
    lostrows = []
    for c, v in bycoc_all.items():
        _g, _r, _lo = coc_lost(v, L)
        if _g is None: continue
        lostrows.append(dict(k=c, city=coc2city.get(c), gmv=round(_g),
            covpct=round(100*_r, 1) if _r is not None else None,
            lost=round(_lo or 0), addpct=round(100*(_lo/_g), 1) if (_lo and _g) else 0))
    lostrows.sort(key=lambda x: -x['lost'])
    # --- tablas por cocina / ciudad ---
    def grp_rows(keyf, keyidx):
        by = defaultdict(list)
        for s in st: by[keyf(s)].append(s)
        out = []
        for k, v in by.items():
            f = [x['fin'] for x in v if x['fin'] is not None]
            rr, rl = rt_lp(keyidx, k)
            out.append(dict(k=k, city=Counter(x['c'] for x in v).most_common(1)[0][0],
                n=len(v), rtwt=None if rr is None else round(rr, 2),
                drt=None if (rr is None or rl is None) else round(rr-rl, 2),
                cov=None if not f else round(100*avg([x['fin']/x['mx'] for x in v if x['fin'] is not None and x.get('mx')]), 1), cov_prop=None,
                dt=[dict(b=x['b'], fin=x['fin'], prop=x['prop']) for x in v]))
        out.sort(key=lambda x: x['cov'] if x['cov'] is not None else 999)
        return out
    cocinas = grp_rows(lambda s: s['k'], 2)
    cities = grp_rows(lambda s: s['c'], 1)
    # --- ventas ---
    def scope_coc(coc): return kind == 'all' or (kind == 'city' and coc2city.get(coc) == name) or (kind == 'op' and coc2op.get(coc) == name)
    ord_coc = {c: coc_tot.get(CUR, {}).get(c, 0) for c in coc_tot.get(CUR, {}) if scope_coc(c)}
    for r in cocinas: r['orders'] = round(ord_coc.get(r['k'], 0))
    for r in cities:  r['orders'] = round(sum(o for c, o in coc_tot.get(CUR, {}).items() if coc2city.get(c) == r['k']))
    # marcas: totales (ALL) o suma de percoc del scope
    if kind == 'all':
        bt = dict(bra_tot.get(CUR, {})); btp = dict(bra_tot.get(PREV, {}))
    else:
        bt = defaultdict(float)
        for coc in cocset:
            if scope_coc(coc):
                for cb, o in percoc.get(coc, []): bt[cb] += o
        bt = dict(bt); btp = {}
    tot_o = sum(bt.values()) or 1
    pareto = []; cum = 0
    for cb, o in sorted(bt.items(), key=lambda x: -x[1]):
        cum += o
        pareto.append(dict(brand=bdisp.get(cb, cb), cb=cb, orders=round(o),
            pct=round(100*o/tot_o, 1), cum=round(100*cum/tot_o, 1)))
    top5 = [p['cb'] for p in pareto[:5]]
    # per_cocina (top 5 marcas por cocina, última semana)
    per_cocina = {}; rank_in = {}
    coc_list = sorted(c for c in cocset if scope_coc(c))
    coc_rt = {c: rt_lp(2, c)[0] for c in coc_list}
    coc_cov = {}; coc_fin = {}
    for c in coc_list:
        v = [s for s in st if s['k'] == c]
        f = [x['fin'] for x in v if x['fin'] is not None]
        coc_cov[c] = round(100*avg([x['fin']/x['mx'] for x in v if x['fin'] is not None and x.get('mx')])) if f else None
        coc_fin[c] = lvl(avg(f)) if f else None
    for coc in coc_list:
        items = sorted(percoc.get(coc, []), key=lambda x: -x[1])
        tt = sum(o for _, o in items) or 1
        per_cocina[coc] = [dict(brand=bdisp.get(cb, cb), orders=round(o), pct=round(100*o/tt, 1)) for cb, o in items[:5]]
        for i, (cb, o) in enumerate(items): rank_in[(coc, cb)] = (i+1, round(o), len(items))
    top5_detail = []
    for cb in top5:
        rows = []
        for coc in coc_list:
            ri = rank_in.get((coc, cb))
            if not ri: continue
            rk, orders, ncb = ri
            rt = coc_rt.get(coc); sz = coc_fin.get(coc); cv = coc_cov.get(coc)
            sev = 'bad' if (rt is not None and rt > 3) or (sz is not None and sz <= 1000) else ('warn' if (rt is not None and rt > 2) or (sz is not None and sz <= 2100) else 'ok')
            rows.append(dict(coc=coc, orders=orders, rank=rk, nbrands=ncb,
                rt=None if rt is None else round(rt, 2), size=sz, cov=cv, sev=sev))
        rows.sort(key=lambda x: -x['orders'])
        wr = [r for r in rows if r['rt'] is not None]
        best = min(wr, key=lambda x: x['rt'])['coc'] if wr else None
        worst = max(wr, key=lambda x: x['rt'])['coc'] if wr else None
        ac = [r['cov'] for r in rows if r['cov'] is not None]
        ar = [r['rt'] for r in rows if r['rt'] is not None]
        oprev = btp.get(cb, 0)
        wow = None if not oprev else round(100*(bt[cb]-oprev)/oprev, 1)
        top5_detail.append(dict(brand=bdisp.get(cb, cb), orders=round(bt.get(cb, 0)), wow=wow,
            avg_cov=round(avg(ac), 1) if ac else None, avg_rt=round(avg(ar), 2) if ar else None,
            best=best, worst=worst, rows=rows))
    n_orders = round(sum(ord_coc.values()))
    return dict(
        kpi=dict(rtwt=None if rtwt is None else round(rtwt, 2),
                 rtwt_lw=None if rtwt_lw is None else round(rtwt_lw, 2),
                 drt=None if (rtwt is None or rtwt_lw is None) else round(rtwt-rtwt_lw, 2),
                 cov=None if cov is None else round(cov, 1),
                 cov_prop=None,
                 n_stores=len(st), n_orders=n_orders, n_brands=len(pareto),
                 n_cocinas=len(cocset), lost=lost, gmv=gmv_now,
                 dist={str(l): dist.get(l, 0) for l in LV}, tot=sum(dist.values()), alerts=alerts),
        stack=dict(labels=[MXW[-2], MXW[-1]], series=stack, tot=stack_tot),
        weekly_rt=wk_net, weekly_cov=weekly_cov, weekly_lost=weekly_lost, lostrows=lostrows,
        cities=cities, cocinas=cocinas,
        pareto=pareto, top5=top5, top5_detail=top5_detail,
        per_cocina=per_cocina, coc_list=coc_list)

data = {'ALL': compute('all', None)}
for c in CITIES:
    r = compute('city', c)
    if r: data[c] = r
for o in OPS:
    r = compute('op', o)
    if r: data['op::'+o] = r

mx_stores = [dict(b=s['b'], k=s['k'], c=s['c'], op=s['op'], fin=s['pw']) for s in stores]
MX = dict(updated='2026-08-24', ideal=IDEAL, LV=LV, cities=CITIES, ops=OPS, weeks=MXW,
          sales_week=sales_week, prev_week=prev_week, data=data, stores=mx_stores)
json.dump(MX, open('mx_data.json', 'w'), ensure_ascii=False)
k = data['ALL']['kpi']
print('MX v2 | semanas', MXW)
print('ciudades', CITIES, '| ops', len(OPS))
print('ALL: RTWT', k['rtwt'], 'vs', k['rtwt_lw'], '(Δ', k['drt'], ') | cob', k['cov'], '% | tiendas', k['n_stores'], '| órdenes', k['n_orders'])
print('weekly_rt', data['ALL']['weekly_rt'])
print('weekly_cov', data['ALL']['weekly_cov'])
print('top5', [bdisp.get(c, c) for c in data['ALL']['top5']])

if '--inject' in sys.argv:
    idx = open('index_mx.html', encoding='utf-8').read()
    payload = '/*MX_DATA_START*/const MX=' + json.dumps(MX, ensure_ascii=False) + ';/*MX_DATA_END*/'
    new = re.sub(r'/\*MX_DATA_START\*/.*?/\*MX_DATA_END\*/', lambda m: payload, idx, count=1, flags=re.S)
    assert new != idx and 'MX_DATA_END' in new
    open('index_mx.html', 'w', encoding='utf-8').write(new)
    print('MX inyectado en index_mx.html')
